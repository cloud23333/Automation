import pandas as pd, pymysql, openpyxl
from openpyxl.styles import PatternFill

origin_path = r"D:\下载\876.xlsx"
jst_path = r"\\Desktop-inv4qoc\图片数据\Temu_半托项目组\倒表格\数据\JST_DATA\Single.xlsx"
out_file = r"C:\Users\Administrator\Documents\Mecrado\Automation\图库的事情\filtered.xlsx"

db_cfg = dict(host="localhost", user="root", password="123456",
              database="temu_gallery", charset="utf8mb4")

def norm(x):
    if x is None or pd.isna(x):
        return ""
    s = str(x).strip()
    return (s[:-2] if s.endswith(".0") else s).lower()

origin = pd.read_excel(origin_path, header=None)
jst = pd.read_excel(jst_path)
sku_set = {norm(v) for v in jst["商品编码"]}

def keep(x):
    s = norm(x)
    return s if s in sku_set else ""

filtered = origin.copy()
filtered.iloc[1:, :] = origin.iloc[1:, :].applymap(keep)

with pymysql.connect(**db_cfg) as conn:
    photo_df = pd.read_sql(
        "SELECT DISTINCT sku_code FROM image_asset WHERE sku_code IS NOT NULL", conn
    )
photo_set = {norm(v) for v in photo_df["sku_code"]}
shot_skus = sku_set & photo_set

for col in filtered.columns:
    data = filtered[col].iloc[1:]
    non_blank = data[data != ""]
    filtered[col] = pd.Series(
        [filtered[col].iloc[0]] + non_blank.tolist() + [""] * (len(filtered) - 1 - len(non_blank))
    )

filtered.to_excel(out_file, index=False, header=False)

wb = openpyxl.load_workbook(out_file)
ws = wb.active
fill = PatternFill(fill_type="solid", start_color="92D050", end_color="92D050")

for row in ws.iter_rows(min_row=2):
    for c in row:
        val = norm(c.value)
        if val and val in shot_skus:
            c.fill = fill

wb.save(out_file)
